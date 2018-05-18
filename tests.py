# -*- coding: utf-8 -*-


def set_to_redis():
    from openpyxl import load_workbook
    import redis
    redis_con = redis.StrictRedis('127.0.0.1', 6379, db=8)
    # Load in the workbook
    file_excel = '/home/meirkhan/Desktop/bot/qa_eng.xlsx'
    wb = load_workbook(file_excel)

    # Get sheet names
    sheet = wb['Лист1']
    data_length = sheet.max_row

    for rownum in range(1, data_length):
        excel_key = sheet.cell(row=rownum, column=1).value.lower()
        excel_value = sheet.cell(row=rownum, column=2).value
        redis_con.set(excel_key, excel_value, 300000)


set_to_redis()

